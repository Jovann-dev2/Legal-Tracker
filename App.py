from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Final
from uuid import uuid4

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ============================================================
# Configuration
# ============================================================
APP_TITLE: Final[str] = "Legal Tracker"
APP_DESCRIPTION: Final[str] = (
    "Upload a source workbook to generate expiry insights by shaft/group, "
    "filter by designation criticality, explore trends, and export production-ready outputs."
)
DEFAULT_YEAR: Final[int] = 2026
VALID_YEAR_RANGE: Final[tuple[int, int]] = (1900, 2100)
MONTH_NUMBERS: Final[list[int]] = list(range(1, 13))
MONTH_NAMES: Final[list[str]] = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]
LEGAL_TYPE_ORDER: Final[list[str]] = [
    "COF Expiry",
    "Work Permit Expiry",
    "Annual Leave Expiry",
]
LEGAL_TYPE_DISPLAY: Final[dict[str, str]] = {
    "COF Expiry": "COF",
    "Work Permit Expiry": "Work permit",
    "Annual Leave Expiry": "Annual leave",
}
CRITICAL_SKILLS_DEFAULT: Final[list[str]] = [
    "Operator Rock Drill Single Handed UG",
    "Operator Winch UG",
    "Operator Loco",
    "Team Leader Production UG",
    "Shift Supervisor Production UG",
    "Miner Stoping",
    "Miner Development",
    "Miner General",
]

st.set_page_config(page_title="Unavailability Insights", layout="wide")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============================================================
# Workbook configuration
# ============================================================
@dataclass(frozen=True)
class SheetConfig:
    legal_type: str
    group_col: str
    date_col: str | None = None
    last_leave_col: str | None = None


SHEET_COLUMN_MAP: Final[dict[str, SheetConfig]] = {
    "COF Register": SheetConfig(
        legal_type="COF Expiry",
        group_col="Group Shaft Name",
        date_col="Next Examination Date",
    ),
    "Work Permits": SheetConfig(
        legal_type="Work Permit Expiry",
        group_col="Group Shaft Name",
        date_col="Permit Expiry Date",
    ),
    "Annual Leave": SheetConfig(
        legal_type="Annual Leave Expiry",
        group_col="Group Shaft Name",
        last_leave_col="Date of Last Leave",
    ),
}


# ============================================================
# Generic helpers
# ============================================================
def normalize_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def validate_year(year: int) -> int:
    min_year, max_year = VALID_YEAR_RANGE
    if min_year <= year <= max_year:
        return year
    return DEFAULT_YEAR


def get_reporting_month_order(start_month: int) -> list[int]:
    return MONTH_NUMBERS[start_month - 1 :] + MONTH_NUMBERS[: start_month - 1]


def get_reporting_window(year: int, start_month: int) -> tuple[pd.Timestamp, pd.Timestamp]:
    start_date = pd.Timestamp(year=year, month=start_month, day=1)
    end_date = start_date + pd.DateOffset(months=12) - pd.Timedelta(days=1)
    return start_date, end_date


def get_reporting_period_label(year: int, start_month: int) -> str:
    start_date, end_date = get_reporting_window(year, start_month)
    return f"{start_date.strftime('%B %Y')} to {end_date.strftime('%B %Y')}"


def find_first_header_row(df: pd.DataFrame, min_non_empty: int = 6, max_scan: int = 50) -> int:
    scan_limit = min(len(df), max_scan)
    for idx in range(scan_limit):
        row = df.iloc[idx]
        non_empty_count = row.map(lambda x: pd.notna(x) and str(x).strip() != "").sum()
        if non_empty_count >= min_non_empty:
            return idx
    return 0


def read_sheet_with_header_detection(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    temp = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine="openpyxl")
    header_row = find_first_header_row(temp)
    header = temp.iloc[header_row].astype(str).tolist()
    data = temp.iloc[header_row + 1 :].copy()
    data.columns = header
    data = data.dropna(how="all")
    return data


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized_targets = [normalize_text(candidate) for candidate in candidates]
    normalized_columns = {column: normalize_text(column) for column in df.columns}

    for column, normalized_column in normalized_columns.items():
        if normalized_column in normalized_targets:
            return column

    for column, normalized_column in normalized_columns.items():
        if any(target in normalized_column for target in normalized_targets):
            return column

    return None


def parse_dates(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    parsed = pd.to_datetime(series, errors="coerce")

    unresolved_mask = parsed.isna()
    numeric_values = pd.to_numeric(series.where(unresolved_mask), errors="coerce")
    excel_serial_mask = unresolved_mask & numeric_values.notna() & numeric_values.between(20000, 60000)

    if excel_serial_mask.any():
        excel_epoch = pd.Timestamp("1899-12-30")
        parsed.loc[excel_serial_mask] = excel_epoch + pd.to_timedelta(
            numeric_values.loc[excel_serial_mask], unit="D"
        )

    return parsed


def classify_sheet(df: pd.DataFrame, sheet_name: str) -> tuple[str | None, dict[str, str]]:
    hard_coded = SHEET_COLUMN_MAP.get(sheet_name)
    if hard_coded:
        required_columns = {
            "group": hard_coded.group_col,
            **({"date": hard_coded.date_col} if hard_coded.date_col else {}),
            **({"last_leave": hard_coded.last_leave_col} if hard_coded.last_leave_col else {}),
        }
        if all(column_name in df.columns for column_name in required_columns.values()):
            return hard_coded.legal_type, required_columns

    group_col = find_column(df, ["group shaft name", "shaft name", "group name", "shaft"])
    cof_col = find_column(df, ["next examination date"])
    permit_col = find_column(df, ["permit expiry date", "permit expiry", "expiry date"])
    leave_col = find_column(df, ["date of last leave", "last leave date", "last leave"])

    if group_col and cof_col:
        return "COF Expiry", {"group": group_col, "date": cof_col}
    if group_col and leave_col:
        return "Annual Leave Expiry", {"group": group_col, "last_leave": leave_col}
    if group_col and permit_col and "permit" in normalize_text(permit_col):
        return "Work Permit Expiry", {"group": group_col, "date": permit_col}

    return None, {}


def apply_designation_filter(
    df: pd.DataFrame,
    filter_mode: str,
    critical_designations: tuple[str, ...],
) -> pd.DataFrame:
    if filter_mode == "Both":
        return df

    designation_col = find_column(df, ["designation"])
    if designation_col is None:
        return df

    normalized_critical = {normalize_text(item) for item in critical_designations if str(item).strip()}
    designation_series = df[designation_col].astype(str).map(normalize_text)

    if filter_mode == "Critical only":
        return df[designation_series.isin(normalized_critical)].copy()
    if filter_mode == "Non-critical only":
        return df[~designation_series.isin(normalized_critical)].copy()

    return df


def ensure_non_empty_group(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    return df[df[group_col].notna() & (df[group_col].astype(str).str.strip() != "")].copy()


def build_monthly_pivot(
    df: pd.DataFrame,
    group_col: str,
    expiry_date: pd.Series,
    legal_type: str,
    year: int,
    start_month: int,
) -> pd.DataFrame:
    valid = df.loc[expiry_date.notna()].copy()
    month_order = get_reporting_month_order(start_month)
    if valid.empty:
        empty = pd.DataFrame(0, index=pd.Index([], name="Shaft"), columns=month_order, dtype=int)
        empty["Legal Type"] = legal_type
        return empty.reset_index().set_index(["Shaft", "Legal Type"])

    valid["_expiry_date"] = expiry_date.loc[expiry_date.notna()]
    valid = ensure_non_empty_group(valid, group_col)

    start_date, end_date = get_reporting_window(year, start_month)
    valid = valid[(valid["_expiry_date"] >= start_date) & (valid["_expiry_date"] <= end_date)].copy()

    if valid.empty:
        empty = pd.DataFrame(0, index=pd.Index([], name="Shaft"), columns=month_order, dtype=int)
        empty["Legal Type"] = legal_type
        return empty.reset_index().set_index(["Shaft", "Legal Type"])

    pivot = (
        valid.groupby([group_col, valid["_expiry_date"].dt.month])
        .size()
        .unstack(fill_value=0)
        .reindex(columns=month_order, fill_value=0)
    )
    pivot.index.name = "Shaft"
    pivot["Legal Type"] = legal_type
    return pivot.reset_index().set_index(["Shaft", "Legal Type"])


def compute_annual_leave_expiry(last_leave_series: pd.Series) -> pd.Series:
    return last_leave_series + pd.DateOffset(months=17)


def build_annual_leave_pivot(
    df: pd.DataFrame,
    group_col: str,
    last_leave_series: pd.Series,
    year: int,
    rollup_month: int,
) -> pd.DataFrame:
    expiry_date = compute_annual_leave_expiry(last_leave_series)
    month_order = get_reporting_month_order(rollup_month)
    valid = df.loc[expiry_date.notna()].copy()

    if valid.empty:
        empty = pd.DataFrame(0, index=pd.Index([], name="Shaft"), columns=month_order, dtype=int)
        empty["Legal Type"] = "Annual Leave Expiry"
        return empty.reset_index().set_index(["Shaft", "Legal Type"])

    valid["_expiry_date"] = expiry_date.loc[expiry_date.notna()]
    valid = ensure_non_empty_group(valid, group_col)

    if valid.empty:
        empty = pd.DataFrame(0, index=pd.Index([], name="Shaft"), columns=month_order, dtype=int)
        empty["Legal Type"] = "Annual Leave Expiry"
        return empty.reset_index().set_index(["Shaft", "Legal Type"])

    shafts = sorted(valid[group_col].astype(str).unique().tolist())
    pivot = pd.DataFrame(0, index=pd.Index(shafts, name="Shaft"), columns=month_order, dtype=int)

    start_date, end_date = get_reporting_window(year, rollup_month)

    # Normal monthly counts for the selected 12-month reporting period
    yearly = valid[(valid["_expiry_date"] >= start_date) & (valid["_expiry_date"] <= end_date)].copy()
    if not yearly.empty:
        monthly_counts = (
            yearly.groupby([group_col, yearly["_expiry_date"].dt.month])
            .size()
            .unstack(fill_value=0)
            .reindex(columns=month_order, fill_value=0)
        )
        for month in month_order:
            if month != rollup_month:
                pivot.loc[monthly_counts.index, month] = monthly_counts[month].astype(int)

    # Roll-up month:
    # Count all annual leave expiries due on or before the end of the selected start month
    rollup_cutoff = pd.Timestamp(year=year, month=rollup_month, day=1) + pd.offsets.MonthEnd(0)
    rollup_counts = (
        valid[valid["_expiry_date"] <= rollup_cutoff]
        .groupby(group_col)
        .size()
        .reindex(pivot.index, fill_value=0)
        .astype(int)
    )
    pivot[rollup_month] = rollup_counts

    pivot["Legal Type"] = "Annual Leave Expiry"
    return pivot.reset_index().set_index(["Shaft", "Legal Type"])


def finalize_result(pivots: list[pd.DataFrame], start_month: int) -> tuple[pd.DataFrame, list[str]]:
    month_order = get_reporting_month_order(start_month)
    columns_out = [*[MONTH_NAMES[month - 1] for month in month_order], "Total"]
    if not pivots:
        empty = pd.DataFrame(columns=["Shaft", "Legal Type", *columns_out])
        return empty, columns_out

    combined = pd.concat(pivots, axis=0).groupby(level=[0, 1]).sum()
    combined = combined.reindex(columns=month_order, fill_value=0)
    combined["Total"] = combined.sum(axis=1)

    month_map = {month: MONTH_NAMES[month - 1] for month in month_order}
    combined = combined.rename(columns=month_map).reset_index()
    combined = combined[["Shaft", "Legal Type", *columns_out]]

    for column in columns_out:
        combined[column] = pd.to_numeric(combined[column], errors="coerce").fillna(0).astype(int)

    return combined.sort_values(["Shaft", "Legal Type"]).reset_index(drop=True), columns_out


# ============================================================
# Cached workbook processing
# ============================================================
@st.cache_data(show_spinner="🔎 Scanning designations…")
def get_all_designations(file_bytes: bytes) -> list[str]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    seen: set[str] = set()

    for sheet_name in excel_file.sheet_names:
        try:
            data = read_sheet_with_header_detection(excel_file, sheet_name)
            designation_col = find_column(data, ["designation"])
            if designation_col is None:
                continue
            values = data[designation_col].dropna().astype(str).str.strip()
            seen.update(value for value in values if value)
        except Exception as exc:
            logger.warning("Failed to scan designations in sheet '%s': %s", sheet_name, exc)

    return sorted(seen)


@st.cache_data(show_spinner="🔄 Reading and aggregating workbook…")
def build_result(
    file_bytes: bytes,
    year: int,
    designation_filter_mode: str,
    critical_designations_selected: tuple[str, ...],
    annual_leave_rollup_month: int,
) -> tuple[pd.DataFrame, list[str], list[str]]:
    excel_file = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    pivots: list[pd.DataFrame] = []
    skipped_sheets: list[str] = []

    for sheet_name in excel_file.sheet_names:
        try:
            data = read_sheet_with_header_detection(excel_file, sheet_name)
            data = apply_designation_filter(data, designation_filter_mode, critical_designations_selected)
            if data.empty:
                continue

            legal_type, column_map = classify_sheet(data, sheet_name)
            if legal_type is None:
                skipped_sheets.append(sheet_name)
                continue

            group_col = column_map["group"]
            if legal_type in {"COF Expiry", "Work Permit Expiry"}:
                expiry_date = parse_dates(data[column_map["date"]])
                pivots.append(
                    build_monthly_pivot(
                        data,
                        group_col,
                        expiry_date,
                        legal_type,
                        year,
                        annual_leave_rollup_month,
                    )
                )
            elif legal_type == "Annual Leave Expiry":
                last_leave_series = parse_dates(data[column_map["last_leave"]])
                pivots.append(
                    build_annual_leave_pivot(
                        data,
                        group_col,
                        last_leave_series,
                        year,
                        annual_leave_rollup_month,
                    )
                )
        except Exception as exc:
            logger.warning("Skipping sheet '%s' because of processing error: %s", sheet_name, exc)
            skipped_sheets.append(sheet_name)

    result, columns_out = finalize_result(pivots, annual_leave_rollup_month)
    return result, columns_out, skipped_sheets


# ============================================================
# Grouping helpers
# ============================================================
def aggregate_by_custom_groups(
    df: pd.DataFrame,
    shaft_to_group_map: dict[str, str],
    month_columns: list[str],
) -> pd.DataFrame:
    if df.empty:
        return df

    aggregated = df.copy()
    aggregated["Shaft"] = aggregated["Shaft"].astype(str)
    aggregated["Shaft"] = aggregated["Shaft"].map(shaft_to_group_map).fillna(aggregated["Shaft"])

    out = (
        aggregated.groupby(["Shaft", "Legal Type"], as_index=False)[month_columns]
        .sum()
        .sort_values(["Shaft", "Legal Type"])
        .reset_index(drop=True)
    )
    out["Total"] = out[month_columns].sum(axis=1).astype(int)

    ordered_columns = ["Shaft", "Legal Type", *month_columns, "Total"]
    return out[ordered_columns]


def get_group_conflicts(group_definitions: list[dict[str, object]]) -> tuple[dict[str, str], list[str]]:
    shaft_to_group: dict[str, str] = {}
    conflicts: set[str] = set()

    for definition in group_definitions:
        group_name = str(definition.get("name") or "Unnamed Group")
        for shaft in definition.get("members", []):
            shaft_str = str(shaft)
            if shaft_str in shaft_to_group and shaft_to_group[shaft_str] != group_name:
                conflicts.add(shaft_str)
            else:
                shaft_to_group[shaft_str] = group_name

    return shaft_to_group, sorted(conflicts)


def init_group_state(shafts: list[str]) -> None:
    shafts_signature = tuple(shafts)
    if st.session_state.get("shafts_signature") != shafts_signature:
        st.session_state.shafts_signature = shafts_signature
        st.session_state.group_defs = []

    st.session_state.setdefault("group_defs", [])


def add_group_definition() -> None:
    st.session_state.group_defs.append(
        {
            "id": str(uuid4()),
            "name": f"Group {len(st.session_state.group_defs) + 1}",
            "members": [],
        }
    )


# ============================================================
# Export helpers
# ============================================================
def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def build_legals_xlsx_bytes(result_df: pd.DataFrame, columns_out: list[str], year: int) -> bytes:
    export_df = result_df.copy()
    if export_df.empty:
        export_df = pd.DataFrame(columns=["Shaft", "Legal Type", *columns_out])

    export_df = export_df[~export_df["Legal Type"].astype(str).str.contains("train", case=False, na=False)].copy()
    month_columns = columns_out[:-1]
    total_column = columns_out[-1]

    for column in [*month_columns, total_column]:
        export_df[column] = pd.to_numeric(export_df[column], errors="coerce").fillna(0).astype(int)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Legals {year}"

    headers = ["Shafts", "Legal type", *month_columns, total_column]
    last_column_index = len(headers)

    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_column_index)
    title_cell = worksheet.cell(row=1, column=2, value=f"Legals {year}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "C3"
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 18
    for column_letter in "CDEFGHIJKLMNO":
        worksheet.column_dimensions[column_letter].width = 12

    thin_side = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
    subtotal_fill = PatternFill(fill_type="solid", fgColor="F4F4F5")

    row_cursor = 3
    shafts_in_order = export_df["Shaft"].astype(str).drop_duplicates().tolist()

    for shaft in shafts_in_order:
        shaft_df = export_df[export_df["Shaft"].astype(str) == shaft].copy()
        first_row_for_shaft = True

        for legal_type in LEGAL_TYPE_ORDER:
            match = shaft_df[shaft_df["Legal Type"] == legal_type]
            values = [0] * len(month_columns)
            total_value = 0

            if not match.empty:
                series = match[month_columns + [total_column]].sum(axis=0)
                values = [int(series[column]) for column in month_columns]
                total_value = int(series[total_column])

            worksheet.cell(row=row_cursor, column=1, value=shaft if first_row_for_shaft else "")
            worksheet.cell(row=row_cursor, column=2, value=LEGAL_TYPE_DISPLAY[legal_type])
            for value_index, value in enumerate([*values, total_value], start=3):
                worksheet.cell(row=row_cursor, column=value_index, value=value)
                worksheet.cell(row=row_cursor, column=value_index).number_format = "0"

            for column_index in range(1, last_column_index + 1):
                worksheet.cell(row=row_cursor, column=column_index).border = border

            first_row_for_shaft = False
            row_cursor += 1

        subtotal_series = shaft_df[month_columns + [total_column]].sum(axis=0)
        subtotal_values = [int(subtotal_series[column]) for column in month_columns]
        subtotal_total = int(subtotal_series[total_column])

        worksheet.cell(row=row_cursor, column=1, value="Total planned Expiries").font = Font(bold=True)
        for value_index, value in enumerate([*subtotal_values, subtotal_total], start=3):
            cell = worksheet.cell(row=row_cursor, column=value_index, value=value)
            cell.font = Font(bold=True)
            cell.number_format = "0"

        for column_index in range(1, last_column_index + 1):
            cell = worksheet.cell(row=row_cursor, column=column_index)
            cell.border = border
            cell.fill = subtotal_fill

        row_cursor += 1

    last_data_row = row_cursor - 1
    if last_data_row >= 3:
        worksheet.conditional_formatting.add(
            f"C3:{worksheet.cell(row=last_data_row, column=last_column_index).coordinate}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFAFA",
                mid_type="percentile",
                mid_value=50,
                mid_color="FCA5A5",
                end_type="max",
                end_color="DC2626",
            ),
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# Chart helpers
# ============================================================

def build_detail_line_chart(df: pd.DataFrame, month_columns: list[str]) -> alt.Chart:
    tidy = df.melt(
        id_vars=["Shaft", "Legal Type", "Total"],
        value_vars=month_columns,
        var_name="Month",
        value_name="Count",
    )

    # Ensure month ordering
    tidy["Month"] = pd.Categorical(tidy["Month"], categories=month_columns, ordered=True)
    tidy["Series"] = tidy["Shaft"].astype(str) + " – " + tidy["Legal Type"].astype(str)

    # Legend-driven multi-select: click legend items to toggle visibility.
    # empty="all" ensures everything shows initially.
    legend_sel = alt.selection_point(
        fields=["Series"],
        bind="legend",
        empty="all",
        toggle=True,     # click toggles series on/off
        clear="dblclick" # double-click clears selection (shows all)
    )

    base = (
        alt.Chart(tidy)
        .encode(
            x=alt.X("Month:O", sort=month_columns, title="Month"),
            y=alt.Y("Count:Q", title="Monthly Count"),
            color=alt.Color("Series:N", title="Series"),
            tooltip=[
                alt.Tooltip("Shaft:N"),
                alt.Tooltip("Legal Type:N"),
                alt.Tooltip("Month:N"),
                alt.Tooltip("Count:Q"),
            ],
        )
    )

    # Make non-selected series faint (or effectively hidden)
    # If you want them fully hidden, set alt.value(0) instead of 0.08
    lines = (
        base.mark_line(point=True)
        .add_params(legend_sel)
        .encode(
            opacity=alt.condition(legend_sel, alt.value(1.0), alt.value(0.08)),
            strokeWidth=alt.condition(legend_sel, alt.value(2.2), alt.value(1.2)),
        )
    )

    # Chart-level formatting similar to your Plotly example
    chart = (
        lines
        .properties(width="container", height=420)
        .configure_legend(
            orient="top",
            direction="horizontal",
            titleOrient="left",
            padding=10,
            labelLimit=260,         # helps with long series names
            symbolStrokeWidth=3
        )
        .configure_axis(
            labelFontSize=11,
            titleFontSize=12,
            grid=True
        )
        .configure_view(stroke=None)  # clean frame
    )

    return chart


def build_totals_by_shaft(df: pd.DataFrame, month_columns: list[str]) -> pd.DataFrame:
    totals = df.groupby("Shaft", as_index=False)[month_columns].sum()
    totals["Total"] = totals[month_columns].sum(axis=1).astype(int)
    return totals.sort_values("Total", ascending=False).reset_index(drop=True)


def build_totals_line_chart(df: pd.DataFrame, month_columns: list[str]) -> alt.Chart:
    tidy = df.melt(id_vars=["Shaft", "Total"], value_vars=month_columns, var_name="Month", value_name="Count")
    tidy["Month"] = pd.Categorical(tidy["Month"], categories=month_columns, ordered=True)

    return (
        alt.Chart(tidy)
        .mark_line(point=True)
        .encode(
            x=alt.X("Month:O", sort=month_columns, title="Month"),
            y=alt.Y("Count:Q", title="Total Expiries"),
            color=alt.Color("Shaft:N", title="Shaft/Group"),
            tooltip=["Shaft", "Month", "Count"],
        )
        .properties(width="container", height=420)
    )


# ============================================================
# UI sections
# ============================================================
def render_sidebar_options() -> tuple[int, int]:
    st.sidebar.header("Reporting Period")
    current_date = pd.Timestamp.now()

    selected_year = st.sidebar.number_input(
        "Year of interest",
        min_value=VALID_YEAR_RANGE[0],
        max_value=VALID_YEAR_RANGE[1],
        value=int(current_date.year),
        step=1,
        help="Choose the starting year for the 12-month reporting window.",
    )

    selected_month_name = st.sidebar.selectbox(
        "Month of interest",
        options=MONTH_NAMES,
        index=current_date.month - 1,
        help=(
            "The report will cover a rolling 12-month period starting from this month "
            "in the selected year."
        ),
    )
    return validate_year(int(selected_year)), MONTH_NAMES.index(selected_month_name) + 1


def render_header(reporting_period_label: str) -> None:
    st.title(APP_TITLE)
    st.caption(APP_DESCRIPTION)
    st.caption(f"📅 Reporting period: {reporting_period_label}")


def render_designation_filters(file_bytes: bytes) -> tuple[str, ...]:
    with st.expander("Designation Filter"):
        dataset_designations = get_all_designations(file_bytes)
        defaults = sorted(set(dataset_designations).intersection(CRITICAL_SKILLS_DEFAULT))
    
        st.caption(
            "Choose which designations count as critical. These will be used for the 'Critical only' output. "
            "The 'Both' output ignores this list."
        )
        selected_critical = st.multiselect(
            "Critical designations",
            options=dataset_designations,
            default=defaults,
            help="These values are treated as the critical designation list for filtering.",
        )
        custom_values = st.text_area(
            "Additional critical designations (optional)",
            placeholder="One per line, or separate with commas / semicolons",
            help="Use this if the workbook contains values not already listed above.",
        )
        if custom_values.strip():
            extras = [item.strip() for item in re.split(r"[,;\n]+", custom_values) if item.strip()]
            selected_critical = sorted(set(selected_critical).union(extras), key=str.lower)
    
        selected_critical_tuple = tuple(sorted(set(selected_critical), key=str.lower))
    return selected_critical_tuple


def render_summary_metrics(df: pd.DataFrame, month_columns: list[str]) -> None:
    if df.empty:
        return

    total_expiries = int(df["Total"].sum())
    unique_shafts = int(df["Shaft"].nunique())
    unique_legal_types = int(df["Legal Type"].nunique())
    month_totals = df[month_columns].sum(axis=0)
    peak_month = month_totals.idxmax() if not month_totals.empty else "N/A"

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total expiries", f"{total_expiries:,}")
    col2.metric("Shafts / groups", unique_shafts)
    col3.metric("Legal types", unique_legal_types)
    col4.metric("Peak month", peak_month)


def render_grouping_section(result: pd.DataFrame, month_columns: list[str], base_filename: str) -> tuple[pd.DataFrame, str]:
    grouping_enabled = st.radio(
        "Would you like to combine shaft names into custom groups?",
        options=["No", "Yes"],
        horizontal=True,
    )

    result_to_show = result.copy()
    output_filename = base_filename

    if grouping_enabled == "No" or result.empty:
        st.session_state["last_exclude_ungrouped"] = False
        return result_to_show, output_filename

    shafts = sorted(result["Shaft"].astype(str).dropna().unique().tolist())
    init_group_state(shafts)

    st.caption("Create one or more groups, assign shafts to each group, and optionally exclude shafts that remain ungrouped.")
    exclude_ungrouped = st.checkbox(
        "Exclude ungrouped shafts",
        value=False,
        help="When enabled, only shafts explicitly assigned to a custom group are included in the aggregated output.",
    )
    st.session_state["last_exclude_ungrouped"] = bool(exclude_ungrouped)

    if not st.session_state.group_defs:
        st.info("No custom groups yet.")
        st.button("➕ Add first group", on_click=add_group_definition)

    group_to_remove_index: int | None = None

    for index, group_def in enumerate(st.session_state.group_defs):
        with st.expander(f"Group {index + 1}", expanded=(index == 0)):
            group_name = st.text_input(
                f"Group name {index + 1}",
                value=group_def["name"],
                key=f"group_name_{group_def['id']}",
            )
            group_members = st.multiselect(
                f"Members for {group_name or group_def['name']}",
                options=shafts,
                default=group_def["members"],
                key=f"group_members_{group_def['id']}",
            )
            group_def["name"] = group_name or f"Group {index + 1}"
            group_def["members"] = group_members

            action_col1, action_col2, _ = st.columns([1, 1, 4])
            with action_col1:
                if st.button("🗑️ Remove", key=f"remove_{group_def['id']}"):
                    group_to_remove_index = index
            with action_col2:
                st.button("➕ Add another", key=f"add_after_{group_def['id']}", on_click=add_group_definition)

    if group_to_remove_index is not None:
        st.session_state.group_defs.pop(group_to_remove_index)
        st.rerun()

    if st.session_state.group_defs:
        st.button("➕ Add another group", on_click=add_group_definition, key="add_group_bottom")

    shaft_to_group, conflicts = get_group_conflicts(st.session_state.group_defs)
    if conflicts:
        st.error(
            "The following shafts are assigned to more than one group: "
            + ", ".join(conflicts)
            + ". Please fix the overlap before proceeding."
        )
        return result_to_show, output_filename

    if not shaft_to_group:
        if exclude_ungrouped:
            st.info("No grouped shafts selected and exclusion is enabled, so the current result is empty.")
            return result.head(0).copy(), output_filename
        st.info("No grouped shafts selected. Showing original results.")
        return result_to_show, output_filename

    data_for_grouping = result.copy()
    if exclude_ungrouped:
        data_for_grouping = data_for_grouping[data_for_grouping["Shaft"].astype(str).isin(shaft_to_group)]
        output_filename = base_filename.replace(".csv", "_grouped_exclusive.csv")
    else:
        output_filename = base_filename.replace(".csv", "_grouped.csv")

    grouped_result = aggregate_by_custom_groups(data_for_grouping, shaft_to_group, month_columns)
    st.success("✅ Custom grouping applied.")
    return grouped_result, output_filename


def render_detailed_analysis(df: pd.DataFrame, month_columns: list[str], reporting_period_label: str) -> None:
    st.subheader(f"Time Series: {reporting_period_label}")
    if df.empty:
        st.info("No data available to display the time series.")
        return

    shafts = sorted(df["Shaft"].astype(str).unique().tolist())
    legal_types = sorted(df["Legal Type"].astype(str).unique().tolist())
    default_shafts = (
        df.groupby("Shaft", as_index=False)["Total"].sum().sort_values("Total", ascending=False)["Shaft"].head(5).tolist()
    )

    control_col, chart_col = st.columns([1, 2], gap="medium")
    with control_col:
        selected_shafts = st.multiselect("Groups/Shafts to plot", options=shafts, default=default_shafts)
        selected_legal_types = st.multiselect("Expiry type(s)", options=legal_types, default=legal_types)

    filtered_df = df[
        df["Shaft"].astype(str).isin(selected_shafts) & df["Legal Type"].astype(str).isin(selected_legal_types)
    ].copy()

    with chart_col:
        if filtered_df.empty:
            st.info("Adjust the filters to display a chart.")
        else:
            st.altair_chart(build_detail_line_chart(filtered_df, month_columns), use_container_width=True)


def render_totals_analysis(df: pd.DataFrame, month_columns: list[str]) -> None:
    if df.empty:
        st.info("No data available to calculate total expiries.")
        return

    totals_df = build_totals_by_shaft(df, month_columns)

    st.subheader("Time Series of Aggregated Expiries (per Shaft)")
    filter_col, chart_col = st.columns([1, 2], gap="medium")
    shaft_options = totals_df["Shaft"].tolist()
    default_selection = shaft_options[:5]

    with filter_col:
        selected_shafts = st.multiselect("Select shaft(s)/group(s)", options=shaft_options, default=default_selection)

    with chart_col:
        if not selected_shafts:
            st.info("Select at least one shaft/group to display the chart.")
        else:
            filtered_totals_df = totals_df[totals_df["Shaft"].isin(selected_shafts)].copy()
            st.altair_chart(build_totals_line_chart(filtered_totals_df, month_columns), use_container_width=True)


def render_downloads_dual(
    df_both: pd.DataFrame,
    df_critical: pd.DataFrame,
    columns_out: list[str],
    year: int,
    csv_filename_both: str,
    csv_filename_critical: str,
) -> None:
    st.subheader("Downloads")
    if df_both.empty and df_critical.empty:
        st.info("No data available to export.")
        return
    else:
        st.write("Below please find nicely-formatted legal tracking data.")

    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("**All Designations**")
        # If a CSV is wanted, the following can be used:
        # st.download_button(
        #     label="Download aggregated CSV",
        #     data=dataframe_to_csv_bytes(df_both),
        #     file_name=csv_filename_both,
        #     mime="text/csv",
        #     use_container_width=True,
        # )
        xlsx_bytes_both = build_legals_xlsx_bytes(df_both, columns_out, year)
        st.download_button(
            label=f"Legals {year} (All Designations)",
            data=xlsx_bytes_both,
            file_name=f"Legals_{year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_right:
        st.markdown("**Critical Designations Only**")
        # If a CSV is wanted, the following can be used:
        # st.download_button(
        #    label="Download aggregated CSV (Critical only)",
        #     data=dataframe_to_csv_bytes(df_critical),
        #     file_name=csv_filename_critical,
        #     mime="text/csv",
        #     use_container_width=True,
        # )
        xlsx_bytes_crit = build_legals_xlsx_bytes(df_critical, columns_out, year)
        st.download_button(
            label=f"Legals {year} (Critical Designations)",
            data=xlsx_bytes_crit,
            file_name=f"Legals_{year}_CriticalOnly.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ============================================================
# Main app
# ============================================================
def main() -> None:
    selected_year, selected_month = render_sidebar_options()
    reporting_period_label = get_reporting_period_label(selected_year, selected_month)
    render_header(reporting_period_label)

    consolidation_tab, analytics_tab = st.tabs(
        ["Consolidation and Workbook Generation", "Analytics"]
    )

    with consolidation_tab:
        uploaded_file = st.file_uploader("📤 Upload the source XLSX workbook", type=["xlsx"])
    
        if uploaded_file is None:
            st.info("Please upload an XLSX file to begin.")
            return
    
        file_bytes = uploaded_file.getvalue()
        selected_critical_designations = render_designation_filters(file_bytes)
    
        result_both, columns_out_both, skipped_sheets_both = build_result(
            file_bytes=file_bytes,
            year=selected_year,
            designation_filter_mode="Both",
            critical_designations_selected=selected_critical_designations,
            annual_leave_rollup_month=selected_month,
        )
    
        result_critical, columns_out_critical, skipped_sheets_critical = build_result(
            file_bytes=file_bytes,
            year=selected_year,
            designation_filter_mode="Critical only",
            critical_designations_selected=selected_critical_designations,
            annual_leave_rollup_month=selected_month,
        )
    
        columns_out = columns_out_both
        month_columns = columns_out[:-1]
        csv_filename_both = f"{selected_year}_expiries_by_shaft.csv"
        csv_filename_critical = f"{selected_year}_expiries_by_shaft_critical_only.csv"
    
        skipped_all = sorted(set(skipped_sheets_both).union(skipped_sheets_critical))
        if skipped_all:
            with st.expander("ℹ️ Workbook processing notes"):
                st.write(
                    "Some sheets were skipped because they did not match the expected layout or could not be processed:"
                )
                st.write(", ".join(skipped_all))
    
        render_summary_metrics(result_both, month_columns)
    
        if result_both.empty:
            st.warning("No matching expiry records were found for the selected reporting period.")
        else:
            with st.expander("Aggregated Results"):
                search_term = st.text_input(
                    "Search shafts / legal types",
                    placeholder="Type part of a shaft or legal type to filter the table",
                ).strip()
        
                filtered_both = result_both.copy()
                filtered_critical = result_critical.copy()
        
                if search_term:
                    mask_both = (
                        filtered_both["Shaft"].astype(str).str.contains(search_term, case=False, na=False)
                        | filtered_both["Legal Type"].astype(str).str.contains(search_term, case=False, na=False)
                    )
                    filtered_both = filtered_both[mask_both].copy()
        
                    mask_crit = (
                        filtered_critical["Shaft"].astype(str).str.contains(search_term, case=False, na=False)
                        | filtered_critical["Legal Type"].astype(str).str.contains(search_term, case=False, na=False)
                    )
                    filtered_critical = filtered_critical[mask_crit].copy()
        
                st.dataframe(filtered_both, use_container_width=True)
    
            with st.expander("Optional Shaft Grouping"):
                result_to_show, csv_filename_both_out = render_grouping_section(filtered_both, month_columns, csv_filename_both)
    
                st.write("### Resulting Dataset")
                st.dataframe(result_to_show, use_container_width=True)

            shaft_to_group, _ = get_group_conflicts(st.session_state.get("group_defs", []))
            exclude_ungrouped_flag = bool(st.session_state.get("last_exclude_ungrouped", False))

            df_critical_for_downloads = filtered_critical.copy()
            if shaft_to_group:
                if exclude_ungrouped_flag:
                    df_critical_for_downloads = df_critical_for_downloads[
                        df_critical_for_downloads["Shaft"].astype(str).isin(shaft_to_group)
                    ].copy()
                    csv_filename_critical_out = csv_filename_critical.replace(".csv", "_grouped_exclusive.csv")
                else:
                    csv_filename_critical_out = csv_filename_critical.replace(".csv", "_grouped.csv")
                df_critical_for_downloads = aggregate_by_custom_groups(
                    df_critical_for_downloads, shaft_to_group, month_columns
                )
            else:
                csv_filename_critical_out = csv_filename_critical

            render_downloads_dual(
                df_both=result_to_show,
                df_critical=df_critical_for_downloads,
                columns_out=columns_out,
                year=selected_year,
                csv_filename_both=csv_filename_both_out,
                csv_filename_critical=csv_filename_critical_out,
            )

        with analytics_tab:
            render_detailed_analysis(result_to_show, month_columns, reporting_period_label)
            render_totals_analysis(result_to_show, month_columns)

if __name__ == "__main__":
    main()
